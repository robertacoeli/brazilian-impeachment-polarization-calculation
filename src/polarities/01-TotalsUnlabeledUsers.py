'''
Calculates the total of posts that an unlabeled user retweeted from each of labeled users of each group ("favoravel" and "contrario").

input: gml files of the retweet networks containing the labeled users (nodes) with their polarities/labels and the unlabeled users (polarity = None).

output: spreadsheet containing the total of retweets for each unlabeled user (total and percentage of retweets in "favoravel"/"coxinhas"; total and percentage of retweets in "contrario"/"petralhas") --> example: ScatterData_UsuariosSemLabel_GranMensal.xlsx (https://drive.google.com/drive/folders/1LivGb9Nddbl2FByLqq6yPezBHxRzfBpT?usp=sharing)

'''

import os
import re
import openpyxl
import networkx as nx

# atualiza o total mensal (agrupa os totais de cada arquivo para um determinado mes)
def updateTotalMensal(scatterMensal, scatterToInsert):
    for (userToInsert, userProps) in scatterToInsert.items():
        if (userToInsert in scatterMensal):
            for (prop, propValue) in userProps.items():
                scatterMensal[userToInsert][prop] += propValue
        else:
            scatterMensal[userToInsert] = userProps.copy()

    return scatterMensal

# obtem os totais para usuarios unlabeled (arquivo diario)
def obterTotaisUsuarios(graph):
    totalDados = dict()

    # for each node (user), checks its edges
    for (user1, userProperty) in graph.nodes(data=True):
        if (userProperty["polaridade"] != "None"):
            continue

        # se não for um usuário marcado (se for um usuario sem label), prossegue a análise
        totalRetweetsMarcados = 0 # total de retweets em usuarios com label (usuarios "marcados")
        numRetweetsCoxinhas = 0 # total de retweets em usuarios "coxinhas" (pro-impeachment)
        numRetweetsPetralhas = 0 # total de retweets em usuarios "petralhas" (anti-impeachment)

        ########################### CALCULO DE TOTAIS (IMPORTANTE) ############################
        # tomo os usuarios predecessores, porque o grafo esta estruturado
        # de forma que os nós recebem as ligações (são um "nó-fim") se
        # tiverem retweetado alguém
        # logo, os predecessores do user1 são aqueles usuários que foram
        # retweetados pelo user1
        #######################################################################################
        for user2 in graph.predecessors(user1):
            user2Polaridade = graph.node[user2]["polaridade"].lower()
            if user2Polaridade == "none":
                continue

            edgeUser2 = graph.edge[user2][user1]
            if "weight" in edgeUser2:
                retweetsUser2 = int(edgeUser2["weight"])
            else:
                retweetsUser2 = int(edgeUser2[0]["weight"])
            totalRetweetsMarcados += retweetsUser2
            if (user2Polaridade == "contrario"):
                numRetweetsPetralhas += retweetsUser2
            else:
                numRetweetsCoxinhas += retweetsUser2

        # ao final, verifica se existiram retweets de usuarios marcados
        # e adiciona ao dict
        if (totalRetweetsMarcados > 0):
            totalDados[user1] = dict()
            totalDados[user1]["NumRetweetsCoxinhas"] = numRetweetsCoxinhas
            totalDados[user1]["PercRetweetsCoxinhas"] = numRetweetsCoxinhas / totalRetweetsMarcados
            totalDados[user1]["NumRetweetsPetralhas"] = numRetweetsPetralhas
            totalDados[user1]["PercRetweetsPetralhas"] = numRetweetsPetralhas / totalRetweetsMarcados

    return totalDados

# agrega os totais de cada mes para os arquivos .gml de grafos, os quais foram gerados com periodicidade mensal
# --> calcula os totais e percentuais de retweets dos usuarios unlabeled por mes
def obterTotalMensal(input_folder):
    outfolder = os.path.dirname(input_folder)
    outfilename = os.path.join(outfolder, "ScatterData_UsuariosSemLabel_GranMensal")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Period")
    ws.cell(row=1, column=2, value="User Id")
    ws.cell(row=1, column=3, value="Number of Retweets - Favoravel (Coxinhas)")
    ws.cell(row=1, column=4, value="Percent of Retweets - Favoravel (Coxinhas)")
    ws.cell(row=1, column=5, value="Number of Retweets - Contrario (Petralhas)")
    ws.cell(row=1, column=6, value="Percent of Retweets - Contrario (Petralhas)")
    rowNumber = 2

    # for each .gml file...
    for root, subdirs, files in os.walk(input_folder):
        for monthFolder in subdirs:
            yearMonth = monthFolder.strip().split("_")
            completeDate = "%s/%s/01" % (yearMonth[0], yearMonth[1])
            totalDadosMes = dict()
            for filename in os.listdir(os.path.join(root, monthFolder)):
                if "ComPosicao.gml" in filename:
                    completeFilename = os.path.join(root, monthFolder, filename)
                    print("Lendo arquivo %s" % completeFilename)
                    graph = nx.read_gml(completeFilename)
                    totalDadosDia = obterTotaisUsuarios(graph) # obtem o total de retweets de cada usuario para cada dia
                    totalDadosMes = updateTotalMensal(totalDadosMes, totalDadosDia) # agrega os dados para verificar os totais por mes

            for (userId, userProps) in totalDadosMes.items():
                ws.cell(row=rowNumber, column=1, value=completeDate)
                ws.cell(row=rowNumber, column=2, value=userId)
                ws.cell(row=rowNumber, column=3, value=userProps["NumRetweetsCoxinhas"])
                ws.cell(row=rowNumber, column=4, value=userProps["PercRetweetsCoxinhas"])
                ws.cell(row=rowNumber, column=5, value=userProps["NumRetweetsPetralhas"])
                ws.cell(row=rowNumber, column=6, value=userProps["PercRetweetsPetralhas"])
                rowNumber += 1

    wb.save(outfilename + ".xlsx")
    print("Finished!")

## MAIN
if __name__ == "__main__":
    # folder containing the .gml files (retweet networks)
    input_folder = "/home/robertacoeli/Documents/Pesquisa/Results/Twitter/Publico/" \
                 "Experimentos_Redes_Retweets_v2/Arquivos_Redes/RedesRetweets_Atualizadas_Maio2018"
    obterTotalMensal(input_folder)